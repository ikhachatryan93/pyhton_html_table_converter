from utilities import Configs
from utilities import create_dir

import data_precessor
from openpyxl import load_workbook

from urllib.request import urljoin

from os import path, sep

# add local modules into PATH
dir_path = path.dirname(path.realpath(__file__))
tmp_data_dir = dir_path + sep + '../output_excel_files'
create_dir(tmp_data_dir)


# def write_in_single_file(urls):
#     tables = []
#     for filename, url in urls.items():
#         print("Extracting tables from {}".format(filename))
#         soup = load_page(url)
#         tables += soup.findAll('table')
#
#    # if tables:
#    #     data_precessor.write_tables_to_excel(tables, 'output', 'www.unknownwebsite.com')


def write_in_multiple_files(urls):
    files = []
    for filename, url in urls.items():
        print("Extracting tables from {}".format(filename))
        try:
            tables = data_precessor.get_interested_tables(url)
        except Exception as e:
            print('Extracting from {} is failed: {}'.format(url, str(e)))
            continue

        # write_tables_to_excel(tables, filename, urljoin(url, '/'))
        if data_precessor.write_tables_to_excel(tables, tmp_data_dir + sep + filename, urljoin(url, '/')):
            files.append(tmp_data_dir + sep + filename + ".xlsx")

    return files


def fix_table_orientation(files):
    for file in files:
        try:
            wb = load_workbook(file)
        except:
            print('Could not find useful data from{}'.format(file))
            continue

        # rotate to horizontal if its vertical
        data_precessor.rotate_to_horizontal(wb)
        wb.save(file)


# def collect_data(files):
#     datas = []
#     for file in files:
#         try:
#             wb = load_workbook(file)
#         except:
#             print('Could not find useful data from{}'.format(file))
#             continue

#         for sheet_name in wb.get_sheet_names():
#             worksheet = wb.get_sheet_by_name(sheet_name)
#             d = data_precessor.Data()
#             path_, name = path.split(file)


def main():
    urls = Configs.get('urls')
    files = write_in_multiple_files(urls)

    print("Processing extracted data...")

    # Comment out if you want to fix table orientation
    # S x        S M L
    # M y   to   x y z
    # L z
    # fix_table_orientation(files)


if __name__ == "__main__":
    main()
